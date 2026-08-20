from __future__ import annotations

import hashlib
import json
import re
import csv
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "docs/cold-start/StoryVerse_seed_integrated_20.xlsx"
OUTPUT_DIR = ROOT / "outputs/storyverse-seed-20"
OUTPUT = OUTPUT_DIR / "StoryVerse_seed_integrated_20_import_ready.xlsx"
OUTPUT_CSV = OUTPUT_DIR / "StoryVerse_seed_integrated_20_import_ready.csv"
CANONICAL_CSV = ROOT / "docs/cold-start/storyverse-seed-stories.csv"
REPORT = OUTPUT_DIR / "validation-report.json"

HEADERS = [
    "external_id",
    "title",
    "body",
    "age",
    "gender",
    "stage",
    "city",
    "latitude",
    "longitude",
    "mood",
    "people",
    "source_note",
    "skip_moderation",
]

CITY_DATA = {
    "北京市": ("北京", 39.9042, 116.4074, "移除行政区后缀并补齐标准城市坐标"),
    "北京": ("北京", 39.9042, 116.4074, "补齐标准城市坐标"),
    "新加坡": ("新加坡", 1.3521, 103.8198, "补齐标准城市坐标"),
    "宁波市": ("宁波", 29.8683, 121.5440, "移除行政区后缀并补齐标准城市坐标"),
    "USA": ("新加坡", 1.3521, 103.8198, "正文明确写明交换地点为 Singapore，修正错误城市"),
    "德黑兰": ("德黑兰", 35.69439, 51.42151, "补齐 Open-Meteo 城市坐标"),
    "上海市": ("上海", 31.2304, 121.4737, "移除行政区后缀并补齐标准城市坐标"),
    "长沙市": ("长沙", 28.2282, 112.9388, "移除行政区后缀并补齐标准城市坐标"),
    "常德市": ("常德", 29.03205, 111.69844, "移除行政区后缀并补齐 Open-Meteo 城市坐标"),
    "深圳市": ("深圳", 22.5431, 114.0579, "移除行政区后缀并补齐标准城市坐标"),
    "遵义市": ("遵义", 27.68667, 106.90722, "移除行政区后缀并补齐 Open-Meteo 城市坐标"),
    "贵阳市": ("贵阳", 26.6470, 106.6302, "移除行政区后缀并补齐标准城市坐标"),
    "永州市": ("永州", 26.42389, 111.61306, "移除行政区后缀并补齐 Open-Meteo 城市坐标"),
    "济南市": ("济南", 36.6512, 117.1201, "移除行政区后缀并补齐标准城市坐标"),
}

ALLOWED_GENDERS = {"男", "女", "其他"}
ALLOWED_STAGES = {"学龄期", "青春期", "成年早期", "成年中期", "老年期"}
ALLOWED_MOODS = {"愤怒", "担心", "失落", "愧疚", "平和自足", "开心幸福", "爱", "自信骄傲"}
ALLOWED_PEOPLE = {"自己", "家人", "恋人", "朋友", "陌生人", "老师", "同事", "其他"}
CJK_PATTERN = re.compile(r"[\u3400-\u9fff\u3040-\u30ff\uac00-\ud7af]")
WORD_PATTERN = re.compile(r"[^\W_]+(?:[’'-][^\W_]+)*", re.UNICODE)


def length_units(text: str) -> tuple[int, str]:
    normalized = text.strip()
    cjk_count = len(CJK_PATTERN.findall(normalized))
    non_cjk = CJK_PATTERN.sub(" ", normalized)
    word_count = len(WORD_PATTERN.findall(non_cjk))
    return cjk_count + word_count, "CJK文字 + 非CJK词" if cjk_count else "非CJK词数"


def body_hash(value: object) -> str:
    return hashlib.sha256(str(value if value is not None else "").encode("utf-8")).hexdigest()


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    target.number_format = source.number_format
    target.protection = copy(source.protection)
    target.alignment = copy(source.alignment)


def style_table_sheet(sheet, widths: dict[str, float], body_height: float = 72) -> None:
    header_fill = PatternFill("solid", fgColor="DCEAF7")
    header_font = Font(name="Aptos", bold=True, color="17324D", size=11)
    body_font = Font(name="Aptos", color="172B3A", size=10)
    thin = Side(style="thin", color="D7E1EA")
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(1, sheet.max_column).column_letter}{sheet.max_row}"
    sheet.row_dimensions[1].height = 27
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color="A8BFD2"))
    for row in sheet.iter_rows(min_row=2):
        sheet.row_dimensions[row[0].row].height = body_height
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.tabColor = "4A90C2"


def add_import_validations(sheet) -> None:
    last_row = max(sheet.max_row, 501)
    validations = [
        (f"D2:D{last_row}", DataValidation(type="whole", operator="between", formula1="1", formula2="120")),
        (f"E2:E{last_row}", DataValidation(type="list", formula1='"男,女,其他"')),
        (f"F2:F{last_row}", DataValidation(type="list", formula1='"学龄期,青春期,成年早期,成年中期,老年期"')),
        (f"H2:H{last_row}", DataValidation(type="decimal", operator="between", formula1="-90", formula2="90", allow_blank=True)),
        (f"I2:I{last_row}", DataValidation(type="decimal", operator="between", formula1="-180", formula2="180", allow_blank=True)),
        (f"J2:J{last_row}", DataValidation(type="list", formula1='"愤怒,担心,失落,愧疚,平和自足,开心幸福,爱,自信骄傲"')),
        (f"M2:M{last_row}", DataValidation(type="list", formula1='"TRUE,FALSE"')),
    ]
    for cell_range, validation in validations:
        validation.error = "该值不符合 StoryVerse 冷启动导入规则。"
        validation.errorTitle = "请检查字段"
        validation.prompt = "请从允许值中选择或填写有效范围。"
        validation.promptTitle = "StoryVerse 导入字段"
        validation.showErrorMessage = True
        validation.showInputMessage = True
        sheet.add_data_validation(validation)
        validation.add(cell_range)


def set_header_comments(sheet) -> None:
    notes = {
        "external_id": "必填且在 seed 故事中唯一；重复编号会跳过。",
        "title": "可留空，由 AI 建议标题。",
        "body": "正文原文；本工作簿中的 20 条正文均与源文件逐字一致。普通用户 100–1500 字/词；授权 seed 允许至 8000 字/词。",
        "age": "故事发生时年龄，1–120 的整数。",
        "gender": "男、女、其他。",
        "stage": "学龄期、青春期、成年早期、成年中期、老年期。",
        "city": "标准城市名，不含“市”等行政区后缀。",
        "latitude": "城市纬度，-90 至 90。",
        "longitude": "城市经度，-180 至 180。",
        "mood": "StoryVerse 规定的主要感受值。",
        "people": "多项必须使用半角 | 分隔。",
        "source_note": "授权或来源说明，仅后台可见。",
        "skip_moderation": "保持 FALSE，所有真实故事继续经过内容审核。",
    }
    for cell in sheet[1]:
        cell.comment = Comment(notes[str(cell.value)], "StoryVerse")


workbook = load_workbook(SOURCE)
source_sheet = workbook["StoryVerse_Seed_Data"]
source_headers = [cell.value for cell in source_sheet[1]]
if source_headers != HEADERS:
    raise ValueError(f"Unexpected headers: {source_headers}")
if source_sheet.max_row != 21:
    raise ValueError(f"Expected 20 source stories, found {source_sheet.max_row - 1}")

source_rows = []
source_hashes = {}
for row_number in range(2, source_sheet.max_row + 1):
    row = {HEADERS[index]: source_sheet.cell(row_number, index + 1).value for index in range(len(HEADERS))}
    external_id = str(row["external_id"] or "").strip()
    if not external_id or external_id in source_hashes:
        raise ValueError(f"Missing or duplicate external_id at row {row_number}")
    source_rows.append(row)
    source_hashes[external_id] = body_hash(row["body"])

original_sheet = workbook.copy_worksheet(source_sheet)
original_sheet.title = "Original_Data"
source_sheet.title = "Import_Ready"
import_sheet = source_sheet

for sheet_name in ["QC_Log", "Summary"]:
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])

change_rows: list[list[object]] = []
qc_rows: list[list[object]] = []
body_hash_rows: list[list[object]] = []

for row_number in range(2, import_sheet.max_row + 1):
    row = {HEADERS[index]: import_sheet.cell(row_number, index + 1).value for index in range(len(HEADERS))}
    external_id = str(row["external_id"] or "").strip()
    body = str(row["body"] if row["body"] is not None else "")

    # Empty titles are intentionally retained so AI can suggest them.
    if row["title"] is None:
        import_sheet.cell(row_number, 2).value = ""

    original_city = str(row["city"] or "").strip()
    if original_city not in CITY_DATA:
        raise ValueError(f"No verified city mapping for {external_id}: {original_city}")
    normalized_city, latitude, longitude, city_reason = CITY_DATA[original_city]
    updates = {"city": normalized_city, "latitude": latitude, "longitude": longitude}
    for field, new_value in updates.items():
        column = HEADERS.index(field) + 1
        old_value = import_sheet.cell(row_number, column).value
        if old_value != new_value:
            import_sheet.cell(row_number, column).value = new_value
            change_rows.append([external_id, field, old_value if old_value is not None else "", new_value, city_reason])

    original_people = str(row["people"] or "").strip()
    normalized_people = "|".join(
        part.strip() for part in re.split(r"[|｜;；、]", original_people) if part.strip()
    )
    if normalized_people != original_people:
        import_sheet.cell(row_number, HEADERS.index("people") + 1).value = normalized_people
        change_rows.append([external_id, "people", original_people, normalized_people, "改为后台实际支持的半角 | 分隔符"])

    import_sheet.cell(row_number, HEADERS.index("skip_moderation") + 1).value = False

    units, length_rule = length_units(body)
    status = "PASS" if 100 <= units <= 1500 else "PASS_SEED_EXCEPTION" if units <= 8000 else "FAIL"
    detail = (
        "符合普通故事 100–1500 字/词规则"
        if status == "PASS"
        else "正文原文完整保留；仅依赖授权冷启动故事 8000 字/词上限"
        if status == "PASS_SEED_EXCEPTION"
        else "超过授权冷启动故事上限"
    )
    qc_rows.append([external_id, length_rule, units, status, detail])
    body_hash_rows.append([external_id, source_hashes[external_id], body_hash(body), "MATCH"])

style_table_sheet(
    import_sheet,
    {"A": 16, "B": 24, "C": 72, "D": 8, "E": 9, "F": 14, "G": 14, "H": 13, "I": 13, "J": 14, "K": 24, "L": 28, "M": 18},
    body_height=84,
)
add_import_validations(import_sheet)
set_header_comments(import_sheet)

style_table_sheet(
    original_sheet,
    {"A": 16, "B": 24, "C": 72, "D": 8, "E": 9, "F": 14, "G": 14, "H": 13, "I": 13, "J": 14, "K": 24, "L": 28, "M": 18},
    body_height=84,
)
original_sheet.sheet_properties.tabColor = "A7B1BA"

qc_sheet = workbook.create_sheet("QC_Log")
qc_sheet.append(["external_id", "length_rule", "length_units", "import_status", "detail"])
for row in qc_rows:
    qc_sheet.append(row)
style_table_sheet(qc_sheet, {"A": 18, "B": 24, "C": 14, "D": 24, "E": 62}, body_height=30)
qc_sheet.conditional_formatting.add(
    f"D2:D{qc_sheet.max_row}",
    FormulaRule(formula=["D2=\"PASS_SEED_EXCEPTION\""], fill=PatternFill("solid", fgColor="FFF1CC")),
)
qc_sheet.conditional_formatting.add(
    f"D2:D{qc_sheet.max_row}",
    FormulaRule(formula=["D2=\"FAIL\""], fill=PatternFill("solid", fgColor="FBD5D5")),
)

summary_sheet = workbook.create_sheet("Summary")
summary_rows = [
    ["Metric", "Value"],
    ["Source stories", len(source_rows)],
    ["Import-ready stories", len(qc_rows)],
    ["Body hashes matched", sum(1 for row in body_hash_rows if row[3] == "MATCH")],
    ["Body text modified", 0],
    ["Standard length PASS", sum(1 for row in qc_rows if row[3] == "PASS")],
    ["Seed length exception", sum(1 for row in qc_rows if row[3] == "PASS_SEED_EXCEPTION")],
    ["Failed rows", sum(1 for row in qc_rows if row[3] == "FAIL")],
    ["English/non-CJK rule", "word count"],
    ["CJK rule", "CJK character count + non-CJK words"],
    ["Ordinary user limit", "100–1500 units"],
    ["Authorised seed limit", "100–8000 units"],
    ["Moderation", "All rows keep skip_moderation=FALSE"],
    ["People delimiter", "ASCII vertical bar |"],
]
for row in summary_rows:
    summary_sheet.append(row)
style_table_sheet(summary_sheet, {"A": 30, "B": 55}, body_height=27)
summary_sheet.sheet_properties.tabColor = "64A879"

change_sheet = workbook.create_sheet("Change_Log")
change_sheet.append(["external_id", "field", "original", "updated", "reason"])
for row in change_rows:
    change_sheet.append(row)
style_table_sheet(change_sheet, {"A": 18, "B": 16, "C": 24, "D": 24, "E": 58}, body_height=35)
change_sheet.sheet_properties.tabColor = "E1A95F"

hash_sheet = workbook.create_sheet("Body_Hash_Audit")
hash_sheet.append(["external_id", "source_sha256", "output_sha256", "result"])
for row in body_hash_rows:
    external_id, source_hash, output_hash, result = row
    hash_sheet.append(
        [
            external_id,
            f"{source_hash[:32]}\n{source_hash[32:]}",
            f"{output_hash[:32]}\n{output_hash[32:]}",
            result,
        ]
    )
style_table_sheet(hash_sheet, {"A": 18, "B": 36, "C": 36, "D": 14}, body_height=40)
hash_sheet.sheet_properties.tabColor = "6E9CC7"

# Keep the import sheet first and make it the active sheet.
workbook._sheets = [
    import_sheet,
    original_sheet,
    qc_sheet,
    summary_sheet,
    change_sheet,
    hash_sheet,
]
workbook.active = 0
workbook.calculation.fullCalcOnLoad = True
workbook.calculation.forceFullCalc = True

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
workbook.save(OUTPUT)

# The administrator page imports CSV. Export the exact Import_Ready grid with
# UTF-8 BOM and full quoting so multiline stories remain byte-for-byte intact.
for csv_path in (OUTPUT_CSV, CANONICAL_CSV):
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.writer(csv_file, quoting=csv.QUOTE_ALL, lineterminator="\n")
        writer.writerow(HEADERS)
        for row_number in range(2, import_sheet.max_row + 1):
            values = [import_sheet.cell(row_number, column).value for column in range(1, len(HEADERS) + 1)]
            values[-1] = "false" if values[-1] is False else "true"
            writer.writerow(["" if value is None else value for value in values])

# Reload the exported workbook and validate values exactly as Excel will read them.
verified = load_workbook(OUTPUT, data_only=False)
verified_import = verified["Import_Ready"]
verified_headers = [cell.value for cell in verified_import[1]]
if verified_headers != HEADERS:
    raise ValueError(f"Exported headers changed: {verified_headers}")
if verified_import.max_row != 21:
    raise ValueError(f"Exported story count is {verified_import.max_row - 1}, expected 20")

verified_ids = []
verified_hashes = {}
validation_errors = []
for row_number in range(2, verified_import.max_row + 1):
    row = {HEADERS[index]: verified_import.cell(row_number, index + 1).value for index in range(len(HEADERS))}
    external_id = str(row["external_id"] or "").strip()
    verified_ids.append(external_id)
    verified_hashes[external_id] = body_hash(row["body"])
    units, _ = length_units(str(row["body"] or ""))
    people = [part.strip() for part in str(row["people"] or "").split("|") if part.strip()]
    checks = {
        "length": 100 <= units <= 8000,
        "age": isinstance(row["age"], int) and 1 <= row["age"] <= 120,
        "gender": row["gender"] in ALLOWED_GENDERS,
        "stage": row["stage"] in ALLOWED_STAGES,
        "city": bool(str(row["city"] or "").strip()),
        "latitude": isinstance(row["latitude"], (int, float)) and -90 <= row["latitude"] <= 90,
        "longitude": isinstance(row["longitude"], (int, float)) and -180 <= row["longitude"] <= 180,
        "mood": row["mood"] in ALLOWED_MOODS,
        "people": bool(people) and all(person in ALLOWED_PEOPLE for person in people),
        "source_note": bool(str(row["source_note"] or "").strip()),
        "skip_moderation": row["skip_moderation"] is False,
        "body_hash": verified_hashes[external_id] == source_hashes.get(external_id),
    }
    failed = [name for name, passed in checks.items() if not passed]
    if failed:
        validation_errors.append({"external_id": external_id, "failed": failed})

if len(set(verified_ids)) != 20:
    validation_errors.append({"external_id": "*", "failed": ["unique_external_id"]})
if validation_errors:
    raise ValueError(f"Validation failed: {validation_errors}")

with OUTPUT_CSV.open("r", encoding="utf-8-sig", newline="") as csv_file:
    csv_rows = list(csv.DictReader(csv_file))
if len(csv_rows) != 20:
    raise ValueError(f"CSV story count is {len(csv_rows)}, expected 20")
for row in csv_rows:
    external_id = row["external_id"].strip()
    if body_hash(row["body"]) != source_hashes.get(external_id):
        raise ValueError(f"CSV body hash mismatch for {external_id}")

report = {
    "source": str(SOURCE),
    "output": str(OUTPUT),
    "csvOutput": str(OUTPUT_CSV),
    "canonicalCsv": str(CANONICAL_CSV),
    "stories": 20,
    "bodyHashMatches": sum(1 for key, value in verified_hashes.items() if value == source_hashes.get(key)),
    "metadataChanges": len(change_rows),
    "standardLengthPass": sum(1 for row in qc_rows if row[3] == "PASS"),
    "seedLengthExceptions": [row[0] for row in qc_rows if row[3] == "PASS_SEED_EXCEPTION"],
    "failedRows": [],
    "sheets": verified.sheetnames,
}
REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
print(json.dumps(report, ensure_ascii=False, indent=2))
